import openpyxl
import random
import string
import hashlib
#STORING PATHS OF THE PRODUCTION, ANONYMOUS DATABASE AND LT
path = "D:\data3.xlsx"
path_lst = "D:\lookup substituted table.xlsx"
path_lt= "D:\lookup table.xlsx"
#LOADING WORKBOOK FOR PRODUCTION DB AND LT
wb_obj = openpyxl.load_workbook(path)
wb_obj_lt= openpyxl.load_workbook(path_lt)
#LOADING ACTIVE SHEET FOR PRODUCTION DB AND LT
sheet_obj = wb_obj.active
sheet_obj_lt=wb_obj_lt.active
#FINDING NO OF ROWS COLS OF PROD DB
row=sheet_obj.max_row
column=sheet_obj.max_column
#DICT FOR LOOKUP TABLE TO ENSURE ONE TO ONE PROPERTY
dict_lt={"Name":"Lookup Name"}

for i in range(1,row+1):
    sheet_obj_lt.cell(row=i,column=1).value = sheet_obj.cell(row=i,column=1).value
for i in range(2,row+1):
    st=sheet_obj_lt.cell(row=i,column=1).value
    print(st)
    dict_lt[st]=hashlib.sha256(st.encode()).hexdigest()

for i in range(2, row + 1):
    sheet_obj_lt.cell(row=i,column=2).value=dict_lt[sheet_obj.cell(row=i,column=1).value]
    sheet_obj.cell(row=i,column=1).value=dict_lt[sheet_obj.cell(row=i,column=1).value]
for i in range(1,row+1):
    for j in range(1,column+1):
        print(sheet_obj.cell(row=i,column=j).value,end=" ")
    print()

wb_obj_lt.save(str(path_lt))

wb_obj.save(str(path_lst))