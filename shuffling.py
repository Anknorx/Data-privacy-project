import openpyxl
import random
import string
path=r"D:\data3.xlsx"
path_shuff=r"D:\shuffled.xlsx"
wb_obj = openpyxl.load_workbook(path)
wb_obj_shuff = openpyxl.load_workbook(path_shuff)
#LOADING THE ACTIVE SHEET:
sheet_obj=wb_obj.active
sheet_obj_shuff=wb_obj_shuff.active
row=sheet_obj.max_row
column=sheet_obj.max_column
ballist=[]
miniballist=[]
#COPYING BANK TABLE TO SHUFFLED TABLE, INITIALLY NO SHUFFLING
for i in range(1,row+1):
    for j in range(1,column+1):
        sheet_obj_shuff.cell(row=i,column=j).value=sheet_obj.cell(row=i,column=j).value
#APPENDING BALANCE VALUES TO THE BALANCE LIST
for i in range(2,row+1):
    ballist.append(sheet_obj_shuff.cell(row=i,column=7).value)

for i in range(2,row+1,100):
    ballist.append(sheet_obj_shuff.cell(row=i+j,column=7).value)
random.shuffle(ballist)



#SHUFFLING THE BALANCE LIST
#random.shuffle(ballist)
#COPYING THE SHUFFLED BALANCE LIST TO shuffled.xlsx
for i in range(2,row+1):
    sheet_obj_shuff.cell(row=i,column=7).value=ballist[i-2]
wb_obj_shuff.save(str(path_shuff))